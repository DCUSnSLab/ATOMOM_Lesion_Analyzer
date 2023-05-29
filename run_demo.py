import os
import numpy as np
from setproctitle import *
import csv
import torch
import torch.onnx
import cv2
import natsort
from run_efficieintNet import Efficient_net
from run_yolo_seg import Yolo_Seg
from ultralytics import YOLO
import openpyxl
from multiprocessing import Process, Queue
def get_images_paths(image_path):
    if (os.path.isfile(image_path)):
        return [image_path]
    elif (os.path.isdir(image_path)):
        file_paths = [x for x in os.listdir(image_path)]
        file_paths = natsort.natsorted(file_paths)
        for i, file in enumerate(file_paths):
            if ('.png' in file or '.jpg' in file or '.JPG' in file):
                pass
            else:
                file_paths[i] = None
        temp_indexes = list(np.where(np.array(file_paths) != None)[0])
        file_paths = [os.path.join(image_path, file_paths[x]) for x in temp_indexes]
        if (len(file_paths) > 0):
            return file_paths
        else:
            raise Exception("No valid image files found, please check dir")
    else:
        raise Exception("image_path is not dir or valid image, please check image_path")

def ordinal_suffix(n):
    suffixes = {1: "st", 2: "nd", 3: "rd"}
    if 10 < n < 20:
        return "%d%s" % (n, "th")
    return "%d%s" % (n, suffixes.get(n % 10, "th"))



class Skin_lesion:
    def __init__(self,ef_configs,yolo_configs,exp=False):

        self.cwd = os.getcwd()
        self.ef_configs = ef_configs
        self.num_ef_models = len(ef_configs)
        self.ef_models, self.ef_weights = self.__load_ef_net(ef_configs=self.ef_configs)

        self.exp = exp
        self.__set_experiment()
        self.yolo_configs = yolo_configs
        self.num_yolo_models = len(yolo_configs)
        self.is_metric_write = False
        self.default_blank = []
        # self.yolo_seg_models = self.__load_yolo(yolo_configs)
    def __set_experiment(self):
        if(self.exp):
            self.exp_outputs = []
            self.exp_wb = openpyxl.Workbook()

            for i in range(self.num_ef_models):
                self.exp_wb.create_sheet(title="eff_"+str(i),index=i)
                self.exp_outputs.append([])
            self.exp_wb.create_sheet(title="ensemble", index=self.num_ef_models)


    def __ensemble_voting(self, ef_results):
        class_names = []
        for i in range(self.num_ef_models):
            class_names += [result[0] for result in ef_results[i]]
        class_names = list(set(class_names))
        ensemble_predictions = {}
        for name in class_names:
            ensemble_predictions[name] = 0.0
        for i, model_result in enumerate(ef_results):
            weight = self.ef_weights[i]
            for j, (class_name, probability) in enumerate(model_result):
                ensemble_predictions[class_name] += weight * probability
        ensemble_predictions = dict(sorted(ensemble_predictions.items(), key=lambda x: x[1], reverse=True))
        return ensemble_predictions
    def __load_ef_net(self,ef_configs):
        models = [None for x in range(self.num_ef_models)]
        ef_weights = [None for x in range(self.num_ef_models)]
        for i, config in enumerate(ef_configs):
            models[i] = Efficient_net(Config=config,device=config.device)
            ef_weights[i] = config.weight
        # NOTE 가중치 정규화, 가중치 합이 1이 아닐시에 정규화를 수행함
        epsilon = 1e-6
        total_weight = sum(ef_weights)
        if (abs(total_weight - 1) < epsilon):
            pass
        else:
            for i in range(len(ef_weights)):
                ef_weights[i] = ef_weights[i] / total_weight
        return models, ef_weights
    def __load_yolo(self,yolo_configs):
        # os.chdir('./segmentation')
        models = [None for x in range(self.num_yolo_models)]
        for i, config in enumerate(yolo_configs):
            models[i] = Yolo_Seg(config=config)
        # os.chdir(self.cwd)
        return models

    def __ef_inference(self, image_info):
        results = [None for x in range(self.num_ef_models)]
        for i, model in enumerate(self.ef_models):
            result = model.inference(image_info=image_info)
            ordinal = ordinal_suffix(i + 1)
            # print(f"{ordinal} model:")
            # print(result)
            results[i] = result
        return results

    def __yolo_seg_inference(self, image_info):
        inference_results = []
        for i, model in enumerate(self.yolo_seg_models):
            inference_results += model.inference(image_info=image_info)
        yolo_inferred_images = []
        yolo_cropped_images = []
        yolo_status = []
        for i, yolo_result in enumerate(inference_results):
            inferred_image, cropped_images, status = yolo_result
            yolo_inferred_images += [inferred_image]
            yolo_cropped_images += [cropped_images]
            yolo_status += [status]
        return yolo_inferred_images, yolo_cropped_images, yolo_status
        # if(self.yolo_configs.verbose):
        #     for i, data in enumerate(inference_results):
        #         for index, cr in enumerate(data[1]):
        #             cv2.imshow(str(index), cr)
        #             cv2.imwrite(str(index) + ".jpg", cr)
        #         cv2.imshow("results", data[0])
        #         cv2.waitKey(0)
        #         cv2.destroyAllWindows()
        # print('*'*50)
    def save_exp(self):
        self.exp_wb.save(str(self.ef_weights) + '.xlsx')
    def __write_metrics_to_excel(self,ws,len_result):
        ws.append([])
        ws.append([])
        temp = ['','True Positive','True Negative','False Positive','False Negative','Precision','Recall','F1-Score','Accuracy','']
        self.default_blank = ['' for x in range(len(temp))]
        temp += ['img_names','Ground Truth']
        for i in range(len_result):
            if( (i)%2 == 0):
                temp.append('predict')
            else:
                temp.append('confidence')
        ws.append(temp)
        return ['',"=COUNTIFS(L:L, \"atopy\", M:M, \"atopy\")","=COUNTIFS(L:L,\"<>Atopy\",M:M,\"<>Atopy\",L:L,\"<>\",M:M,\"<>\")-1", "=COUNTIFS(L:L, \"<>Atopy\", M:M, \"Atopy\")","=COUNTIFS(L:L, \"Atopy\", M:M, \"<>Atopy\")","=B4/(B4+D4)","=B4/(B4+E4)","=2*(F4*G4)/(F4+G4)","=(B4+C4)/(B4+C4+D4+E4)",'']

        # ws.append
        pass

    def __write_to_sheet(self, ws, img_name, result, is_metric_write):
        result = [x for pair in result for x in pair]
        if is_metric_write:
            metric = self.__write_metrics_to_excel(ws, len(result))
            ws.append(metric + [img_name, img_name.split("_")[0]] + result)
        else:
            ws.append(self.default_blank + [img_name, img_name.split("_")[0]] + result)

    def __write_excel(self, img_name, ef_results, ensemble_predictions):
        is_metric_write = not self.is_metric_write
        if is_metric_write:
            self.is_metric_write = True
        for i, result in enumerate(ef_results):
            ws = self.exp_wb[self.exp_wb.sheetnames[i]]
            self.__write_to_sheet(ws, img_name, result, is_metric_write)

        ws = self.exp_wb['ensemble']
        self.__write_to_sheet(ws, img_name, ensemble_predictions, is_metric_write)

    def inference(self,image_path):
        img = cv2.imread(image_path)
        ef_results = self.__ef_inference(img)
        ensemble_predictions = self.__ensemble_voting(ef_results=ef_results)
        # print('*'*50)
        # print(ensemble_predictions)
        # print(list(ensemble_predictions.items()))
        if(self.exp):
            self.__write_excel(img_name=os.path.basename(image_path).split('.')[0], ef_results=ef_results,
                               ensemble_predictions=list(ensemble_predictions.items()))
        # exit()
        # yolo_inferred_images, yolo_cropped_images, yolo_status = self.__yolo_seg_inference(img)





if __name__ == '__main__':
    setproctitle('lesion')

    class Config_41():
        weight = 0.5
        device = 0
        verbose = False
        topk = 5
        model_path = os.path.join(os.getcwd(),"classification/efficientnet_models/41.pt")
        model_name = 'efficientnet-b0'
        class_names = {
        "000": "normal_skin",
        "001": "atopy",
        "002": "prurigo",
        "003": "scar",
        "004": "psoriasis",
        "005": "varicella",
        "006": "nummular_eczema",
        "007": "ota_like_melanosis",
        "008": "becker_nevus",
        "009": "pyogenic_granuloma",
        "010": "acne",
        "011": "salmon_patches",
        "012": "dermatophytosis",
        "013": "wart",
        "014": "impetigo",
        "015": "vitiligo",
        "016": "ingrowing_nails",
        "017": "congenital_melanocytic_nevus",
        "018": "keloid",
        "019": "epidermal_cyst",
        "020": "insect_bite",
        "021": "molluscum_contagiosum",
        "022": "pityriasis_versicolor",
        "023": "melanonychia",
        "024": "alopecia_areata",
        "025": "epidermal_nevus",
        "026": "herpes_simplex",
        "027": "urticaria",
        "028": "nevus_depigmentosus",
        "029": "lichen_striatus",
        "030": "mongolian_spot_and_ectopic_mongolian_spot",
        "031": "capillary_malformation",
        "032": "pityriasis_lichenoides_chronica",
        "033": "infantile_hemangioma",
        "034": "mastocytoma",
        "035": "nevus_sebaceous",
        "036": "onychomycosis",
        "037": "milk_coffee_nevus",
        "038": "nail_dystrophy",
        "039": "melanocytic_nevus",
        "040": "juvenile_xanthogranuloma",
        }
    class Config_41_min(Config_41):
        weight = 0.1
        model_path = os.path.join(os.getcwd(),"classification/efficientnet_models/dp05_dc05.pt")
    class Config_6_min(Config_41):
        weight = 0.3
        device = 1
        model_path = "classification/efficientnet_models/min_b0_6.pt"
        class_names = {
        "000": "normal_skin",
        "001": "atopy",
        "002": "psoriasis",
        "003": "acne",
        "004": "epidermal_cyst",
        "005": "varicella",
        }
    class Config_5(Config_41):
        weight = 0.1
        device = 1
        model_path = "classification/efficientnet_models//classify_5.pt"
        model_name = 'efficientnet-b7'
        class_names = {
            "000": "atopy",
            "001": "seborrheic dermatitis",
            "002": "psoriasis",
            "003": "rosacea",
            "004": "acne",
        }

    class Config_yolo():
        model_path = "segmentation/yolo_models/best_n.pt"
        model_names = None
        display = False
        save_path = None
        verbose = False
        device = 1
        label = True
        bbox = True
        segmentation = True
        file_paths = None


    ef_configs = [Config_41, Config_41_min,Config_6_min,Config_5]
    yolo_configs = [Config_yolo]
    skin_lesion = Skin_lesion(ef_configs=ef_configs,yolo_configs=yolo_configs,exp=True)
    image_path = "test_data/atomom_test_images_samples/"
    image_path_list = get_images_paths(image_path)
    output = []
    for i, image_path in enumerate(image_path_list):
        skin_lesion.inference(image_path=image_path)
        if(i>10):
            break
    if(skin_lesion.exp):
        skin_lesion.save_exp()
        # output.append([file_name, class_name, confidence])
    # Write the output to a CSV file
    # with open('./experiment_results/output.csv', 'w', newline='') as file:
    #     writer = csv.writer(file)
    #     writer.writerows(output)
